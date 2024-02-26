import openpyxl
from openpyxl import load_workbook
import calendar
wb = load_workbook("mad20230603.xlsx")  # Work Book
s1 = wb['63']  # Work Sheet

count = [[0]*65 for t in range(55)]
count2=[[0]*65 for t in range(26)]

for x in range(0,s1.max_row):
            for y in range(0,s1.max_column):
                count[x][y]=s1.cell(x+1,y+1).value

num1=0
num2=0
for x in range(2,len(count)):
    for y in range(5,len(count[0])):
        if count[x][y]!= 0:
            num1=num1+count[x][y]
            num2=num2+1
    ave=num1/num2
    day=count[x][2]*2-1
    if count[x][4]=="A":
        count2[count[x][3]+2][day]=ave
    elif count[x][4]=="B":
        count2[count[x][3]+2][day+1]=ave
for i in range(2,26):
      count2[i][0]=i-2

x, y = calendar.monthrange(count[2][0] , count[2][1])
k=1
for j in range(1,y+1):
      count2[0][k]=str(count[2][1])+"/"+str(j)
      count2[0][k+1]=str(count[2][1])+"/"+str(j)
      k=k+2
      print(k)
count2[0][0]="Date"
count2[1][0]="channal"
for c in range(1,65):
    if c%2==0:
        count2[1][c]="B"
    else:
        count2[1][c]="A"


wb2 = openpyxl.Workbook()
sheet = wb2.create_sheet("all")
for x in range(0,len(count2)):
            for y in range(0,len(count2[0])):
                sheet.cell(x+1,y+1).value=count2[x][y]
name="test.xlsx"
wb2.save(name)
