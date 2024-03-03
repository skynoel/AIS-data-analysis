import re
import openpyxl
import os
from openpyxl.styles import PatternFill

os.system("cls")

dwb = openpyxl.load_workbook("mad202304.xlsx")  # Work Book
s1 = dwb['all']  # Work Sheet
wb = openpyxl.Workbook()


data =[[0]*65 for t in range(26)]
result=[[0]*65 for t in range(26)]

for x in range(0,s1.max_row):
            for y in range(0,s1.max_column):
                data[x][y]=s1.cell(x+1,y+1).value
for x in range(0,s1.max_row):
            for y in range(0,s1.max_column):
                result[x][y]=s1.cell(x+1,y+1).value
yave=130
total=0
count=0

for x in range(2,len(data)):
    for y in range(1,len(data[0])):
        if data[x][y]!=0:     
            total=total+data[x][y]
            count=count+1
mave=total/count

if yave<=mave:
    for y in range(1,len(data[0])):
        for x in range(2,len(data)):
            if x<3:
                if (data[x][y]+data[x+1][y])/2 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0
            elif x==len(data)-1:
                if (data[x][y]+data[x-1][y])/2 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0
            else:
                if (data[x][y]+data[x+1][y]+data[x-1][y])/3 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0
else:
    for y in range(1,len(data[0])):
        for x in range(2,len(data)):
            if x<3:
                if (data[x][y]+data[x+1][y])/2 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0
            elif x==len(data)-1:
                if (data[x][y]+data[x-1][y])/2 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0
            else:
                if (data[x][y]+data[x+1][y]+data[x-1][y])/3 >=mave:
                    result[x][y]=1
                else:
                    result[x][y]=0

sheet = wb.create_sheet("patd")
fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
for x in range(0,len(result)):
            for y in range(0,len(result[0])):
                sheet.cell(x+1,y+1).value=result[x][y]
                
for x in range(2,sheet.max_row):
            for y in range(1,sheet.max_column):
                if sheet.cell(x+1,y+1).value==1:
                    sheet.cell(x+1,y+1).fill=fill

sheet2 = wb.create_sheet("patd2")
for x in range(0,sheet.max_row):
    for y in range(0,sheet.max_column):
        sheet2.cell(y+2,x+1).value=sheet.cell(x+1,y+1).value
for x in range(1,sheet2.max_row):
    for y in range(2,sheet2.max_column):
        if sheet2.cell(x+1,y+1).value==1:
            sheet2.cell(x+1,y+1).fill=fill
for x in range(0,sheet2.max_row):
    for col in sheet2.columns:
            col_name = re.findall('\w\d', str(col[0]))
            col_name = col_name[1]
            col_name = re.findall('\w', str(col_name))[0]
            sheet2.column_dimensions[col_name].width =9
            sheet2.row_dimensions[x+2].height =50
sheet2.cell(1,1).value="月平均值"
sheet2.cell(1,2).value=mave
del wb["Sheet"]
name="patd202304.xlsx"
wb.save(name)
print("done")