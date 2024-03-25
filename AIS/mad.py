import numpy
import openpyxl
import pandas as pd
import os
from tqdm import tqdm
import calendar
import matplotlib.pyplot as plt

def avperm(folder_path,name):
    wb = openpyxl.Workbook()

    count2=[[0]*65 for t in range(26)]

    folders = os.listdir(folder_path)
    for folder in tqdm(folders):
        folder_path1 = os.path.join(folder_path, folder)
        if os.path.isdir(folder_path1):
            count = [[0]*65 for t in range(4)]
            count[0][0]="RMC_DATE_YEAR"
            count[0][1]="RMC_DATE_MON"
            count[0][2]="RMC_DATE_DAY"
            count[0][3]="RMC_TAKEN_AT_HOUR"
            count[0][4]="CHANNEL"
            for y in range(5):
                count[1][y]=" "
            for y in range(60):
                count[1][y+5]=y
            for y in range(60):
                count[0][y+5]=" "
            row=2
            rowchange=30
            r2=40
            items = os.listdir(folder_path1)
            for item in tqdm(items):
                item_path = os.path.join(folder_path1, item)
                if os.path.isfile(item_path):
                    print("檔案:", item)
                    df=pd.read_excel(item_path)
                    for i in tqdm(range(df.shape[0])):
                        data=df.iloc[i]
                        a=isinstance(data.iloc[3],numpy.int64)
                        b=isinstance(data.iloc[2],numpy.int64)
                        if a and b:
                            if rowchange<data.iloc[3]:
                                row=row+2
                                count.append([0]*65)
                                count.append([0]*65)
                                #print(row)
                                rowchange=data.iloc[3]
                            elif r2<data.iloc[2]:
                                row=row+2
                                count.append([0]*65)
                                count.append([0]*65)
                                #print(row)
                                r2=data.iloc[2]
                            if data.iloc[8] == "A":
                                for k in range(0,4):
                                    count[row][k]=int(data.iloc[k])
                                count[row][4]="A"
                                count[row][data.iloc[4]+5]=int(count[row][data.iloc[4]+5])+1
                                rowchange=data.iloc[3]
                                r2=data.iloc[2]
                            elif data.iloc[8] == "B":
                                for k in range(0,4):
                                    count[row+1][k]=int(data.iloc[k])
                                count[row+1][4]="B"
                                count[row+1][data.iloc[4]+5]=int(count[row+1][data.iloc[4]+5])+1
                                rowchange=data.iloc[3]
                                r2=data.iloc[2]
            sheetname=str(data.iloc[0])+str(data.iloc[1])+str(data.iloc[2])
            sheet = wb.create_sheet(sheetname)

            for x in range(0,len(count)):
                for y in range(0,len(count[0])):
                    sheet.cell(x+1,y+1).value=count[x][y]
            num1=0
            num2=0
            for x in range(2,len(count)):
                for y in range(5,len(count[0])):
                    if count[x][y]!= 0:
                        num1=num1+count[x][y]
                        num2=num2+1
                if num2!=0:
                    ave=num1/num2
                    day=count[x][2]*2-1
                    if count[x][4]=="A":
                        count2[count[x][3]+2][day]=ave
                    elif count[x][4]=="B":
                        count2[count[x][3]+2][day+1]=ave

    for i in range(2,26):
        count2[i][0]=i-2            
    x, y = calendar.monthrange(2023 , 4)
    k=1
    for j in range(1,y+1):
        count2[0][k]=str(4)+"/"+str(j)
        count2[0][k+1]=str(4)+"/"+str(j)
        k=k+2
    count2[0][0]="Date"
    count2[1][0]="channal"
    for c in range(1,65):
        if c%2==0:
            count2[1][c]="B"
        else:
            count2[1][c]="A"

    sheet = wb.create_sheet("all")
    for x in range(0,len(count2)):
                for y in range(0,len(count2[0])):
                    sheet.cell(x+1,y+1).value=count2[x][y]

    wb.save(name)  
    print("done")
    return name


def graph(file):
    data=openpyxl.load_workbook(file)
    sheet=data['all']
    x=[]
    y=[]

    for k in range(2,sheet.max_column):
        for i in range(3,sheet.max_row):
            A=sheet.cell(i,1).value
            x.append(A)
        for j in range(3,sheet.max_row):
            B=sheet.cell(j,k).value
            y.append(B)
        plt.plot(x, y, marker='.',color='red',linewidth=0.1,label=str(sheet.cell(1,k).value)+sheet.cell(2,k).value)
    plt.xlabel('hour')
    plt.ylabel('mad')
    plt.title('madfile')
    plt.savefig(file.rstrip(".xlsx")+".png") 
    plt.show()
    print("done")