import re
import openpyxl
import os
from openpyxl.styles import PatternFill

os.system("cls")
 
dwb = openpyxl.load_workbook("patd202304.xlsx")  # Work Book
s1 = dwb['patd2']  # Work Sheet
wb = openpyxl.Workbook()


data =[[0]*26 for t in range(66)]
result =[[0]*26 for t in range(66)]

for x in range(0,s1.max_row):
            for y in range(0,s1.max_column):
                data[x][y]=s1.cell(x+1,y+1).value
for x in range(0,s1.max_row):
            for y in range(0,s1.max_column):
                result[x][y]=s1.cell(x+1,y+1).value

for x in range(2,len(data)):
    for y in range(4,len(data[0])-2):
        if (data[x][y+2]+data[x][y+1]+data[x][y]+data[x][y-1]+data[x][y-2])>=3:
                if result[x][y] !=1:
                    result[x][y]=2
for x in range(2,len(data)):
    for y in range(2,len(data[0])):
        if (data[x][1]=="A"):
               if result[x][y] ==1 or result[x][y]==2:
                    if result[x+1][y] !=1:
                        result[x+1][y]=2
        if (data[x][1]=="B"):
               if result[x][y] ==1 or result[x][y]==2:
                    if result[x-1][y] !=1:
                        result[x-1][y]=2

sheet = wb.create_sheet("chr_patd")
for x in range(0,len(result)):
            for y in range(0,len(result[0])):
                sheet.cell(x+1,y+1).value=result[x][y]


fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
fill2=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")

for x in range(2,sheet.max_row):
            for y in range(1,sheet.max_column):
                if sheet.cell(x+1,y+1).value==1:
                    sheet.cell(x+1,y+1).fill=fill
                if sheet.cell(x+1,y+1).value==2:
                    sheet.cell(x+1,y+1).fill=fill2













del wb["Sheet"]
name="chr_patd202304.xlsx"
wb.save(name)
print("done")






